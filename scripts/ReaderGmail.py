import openpyxl
from random import randint, choice
import string

def generate_random_email():
    domains = ["gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com"]
    username = ''.join(choice(string.ascii_lowercase) for _ in range(8))
    domain = choice(domains)
    return f"{username}@{domain}"

def generate_random_password():
    password = ''.join(choice(string.ascii_letters + string.digits) for _ in range(10))
    return password

def generate_test_data(rows, cols):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Generate headers
    sheet.cell(row=1, column=1, value="Gmail")
    sheet.cell(row=1, column=2, value="Password")
    for col in range(3, cols + 1):
        sheet.cell(row=1, column=col, value=f"Column {col}")

    # Generate data
    for row in range(2, rows + 2):
        sheet.cell(row=row, column=1, value=generate_random_email())
        sheet.cell(row=row, column=2, value=generate_random_password())
        for col in range(3, cols + 1):
            # Generating random integer data for simplicity
            sheet.cell(row=row, column=col, value=randint(1, 100))

    return wb

def save_excel_file(workbook, filepath):
    workbook.save(filepath)

# Example usage: Generate a 100x5 Excel file and save to custom directory
rows = 100  # Number of rows
cols = 5    # Number of columns
custom_directory = "./data"
filename = "test_data.xlsx"
file_path = custom_directory + filename
excel_file = generate_test_data(rows, cols)
save_excel_file(excel_file, file_path)
print(f"Excel file '{filename}' saved successfully to '{custom_directory}'.")
